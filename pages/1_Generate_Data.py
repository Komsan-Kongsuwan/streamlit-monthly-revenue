# pages/2_Customer_Report.py
import streamlit as st
import pandas as pd
import numpy as np
import io
import calendar
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# --- Warehouse mapping (same as your original) ---
warehouse_site_map = {
    "WAREHOUSE ( ACW )": "ACW",
    "WAREHOUSE ( TAC )": "TAC",
    "WAREHOUSE ( WELLGROW - WGR )": "WGR",
    "WAREHOUSE ( PHRA PRADAENG )": "PPD",
    "WAREHOUSE ( SSW )": "SSW",
    "WAREHOUSE ( BANGPLEE - WBP )": "BP1",
    "WAREHOUSE ( BANGPLEE - WH.A )": "BPA",
    "WAREHOUSE ( BANGPLEE - WH.B )": "BPB",
    "WAREHOUSE ( BANGPLEE - WH.C )": "BPC",
    "WAREHOUSE ( BANGPLEE - WH.D )": "BPD",
    "WAREHOUSE ( ST9 )": "ST9",
    "WAREHOUSE ( NEXT GEN )": "NeG",
    "WAREHOUSE ( NeG )": "NeG",
    "WAREHOUSE ( ST11 )": "ST11",
    "WAREHOUSE ( STF )": "STF",
    "WAREHOUSE ( TBN )": "TBN1",
    "WAREHOUSE (P304 )": "P304",
    "WAREHOUSE (UAF )": "UAF",
    "WAREHOUSE ( HMW )": "HMW",
    "WAREHOUSE ( TBN2 )": "TBN2",
    "WAREHOUSE (BN20 )": "BN20",
    "WAREHOUSE (OTHER)": "OTHER",
    "WAREHOUSE (ZC )": "ZC"
}

def extract_site(warehouse_desc):
    if pd.isna(warehouse_desc):
        return "Unknown"
    return warehouse_site_map.get(str(warehouse_desc).strip(), "Unknown")


def find_columns(df):
    """
    Find the CUSTOMER column and the REVENUE column dynamically.
    Returns column labels (not integer positions).
    Raises ValueError if not found.
    """
    # Convert every cell to uppercase string for matching
    df_str = df.astype(str).applymap(lambda x: x.strip().upper() if pd.notna(x) else "")

    customer_col = None
    amount_col = None

    # Search rows for CUSTOMER and REVENUE keywords
    for r in range(len(df_str)):
        row = df_str.iloc[r]
        # Find CUSTOMER column in this row (word boundary)
        cust_matches = row[row.str.contains(r'\bCUSTOMER\b', regex=True)]
        if not cust_matches.empty and customer_col is None:
            customer_col = cust_matches.index[0]
        # Find REVENUE column in this row (word boundary)
        rev_matches = row[row.str.contains(r'\bREVENUE\b', regex=True)]
        if not rev_matches.empty and amount_col is None:
            # prefer exact 'REVENUE' equality if present in that row, otherwise first match
            exact = row[row == "REVENUE"]
            if not exact.empty:
                amount_col = exact.index[0]
            else:
                amount_col = rev_matches.index[0]

        if customer_col is not None and amount_col is not None:
            break

    if customer_col is None:
        raise ValueError("Could not find 'CUSTOMER' column.")
    if amount_col is None:
        raise ValueError("Could not find 'REVENUE' column.")

    return customer_col, amount_col


# ---------------- Streamlit UI ----------------
st.set_page_config(page_title="Customer Monthly Report", layout="wide")
st.title("📊 Customer Monthly Report")

st.write("Upload multiple customer report Excel files (filename must include YYYYMM).")
uploaded_files = st.file_uploader("Upload .xlsx files", type="xlsx", accept_multiple_files=True)

if not uploaded_files:
    st.info("Please upload one or more monthly Excel files (e.g. 202401.xlsx).")
    st.stop()

# progress and status placeholders
progress = st.progress(0)
status = st.empty()

all_data = []
total_files = len(uploaded_files)

for idx, upfile in enumerate(uploaded_files):
    try:
        status.text(f"Processing {upfile.name} ({idx+1}/{total_files})")
        # read first sheet
        xls = pd.ExcelFile(upfile)
        sheet_name = xls.sheet_names[0]
        df = pd.read_excel(xls, sheet_name=sheet_name, dtype=object)
    except Exception as e:
        status.text(f"Skipping {upfile.name} due to read error: {e}")
        progress.progress(int(((idx+1)/total_files) * 100))
        continue

    # try to find columns
    try:
        customer_col, amount_col = find_columns(df)
    except Exception as e:
        status.text(f"Skipping {upfile.name}: {e}")
        progress.progress(int(((idx+1)/total_files) * 100))
        continue

    # locate warehouse rows in first column
    first_col = df.columns[0]
    warehouse_mask = df[first_col].astype(str).str.contains("WAREHOUSE", case=False, na=False)
    warehouse_rows = df[warehouse_mask].index.tolist()

    # assign site from the warehouse header rows
    for i, row_index in enumerate(warehouse_rows):
        site = extract_site(df.iloc[row_index, 0])
        next_row_index = warehouse_rows[i + 1] if i + 1 < len(warehouse_rows) else len(df)
        # assign site to rows between (row_index+1) and (next_row_index-1)
        df.loc[row_index + 1: next_row_index - 1, 'Site'] = site

    # pick relevant columns and drop rows with missing keys
    try:
        data = df[[customer_col, amount_col, 'Site']].dropna(subset=[customer_col, amount_col, 'Site'])
    except Exception as e:
        status.text(f"Skipping {upfile.name}: cannot extract necessary columns: {e}")
        progress.progress(int(((idx+1)/total_files) * 100))
        continue

    # remove rows where customer header text appears in the column
    data = data[~data.iloc[:,0].astype(str).str.contains('CUSTOMER', case=False, na=False)]

    # parse date from filename - expect 'YYYYMM' before extension
    try:
        date_str = upfile.name.split('.')[0]
        date = datetime.strptime(date_str, '%Y%m')
    except Exception:
        # fallback: if not parseable, use today's date
        date = datetime.now()

    data['Date'] = date.replace(day=1)
    data['Month'] = date.month
    data['Year'] = date.year

    # rename columns to consistent names
    data.columns = ['Customer', 'Amount', 'Site', 'Date', 'Month', 'Year']

    all_data.append(data)

    # update progress
    progress.progress(int(((idx+1)/total_files) * 100))

# After processing all files
if not all_data:
    status.text("No valid data processed from uploaded files.")
    st.stop()

df_final = pd.concat(all_data, ignore_index=True)

# aggregate all-site SDCT row
df_all_site = df_final.groupby(['Year', 'Month'], as_index=False).agg({'Date': 'first', 'Amount': 'sum'})
df_all_site['Site'] = 'SDCT'
df_all_site['Customer'] = 'All Customer'
df_final = pd.concat([df_final, df_all_site], ignore_index=True)

# custom site order, then ensure categorical ordering
custom_order = ['SDCT', 'ACW', 'BPA', 'BPB', 'BPC', 'BPD', 'BN20']
all_sites = df_final['Site'].dropna().unique().tolist()
full_order = custom_order + [x for x in all_sites if x not in custom_order]
site_order = pd.CategoricalDtype(categories=full_order, ordered=True)
df_final['Site'] = df_final['Site'].astype(site_order)
df_final = df_final.sort_values(by=['Site', 'Year'])

# format Month as two-digit string
df_final['Month'] = df_final['Month'].astype(int).astype(str).str.zfill(2)

# pivot table (Site, Customer, Year) vs months
pivot_df = pd.pivot_table(
    df_final,
    index=['Site', 'Customer', 'Year'],
    columns='Month',
    values='Amount',
    aggfunc='sum',
    fill_value=0,
    observed=False
)

pivot_df['Grand Total'] = pivot_df.sum(axis=1)
pivot_df = pivot_df.reset_index()

# drop all-zero rows
pivot_df = pivot_df[pivot_df['Grand Total'] != 0]

# rename month columns
month_map = {f"{i:02d}": calendar.month_abbr[i] for i in range(1, 13)}
pivot_df = pivot_df.rename(columns=month_map)

# ensure final columns ordering
month_names = [calendar.month_abbr[i] for i in range(1, 13)]
final_columns = ['Site', 'Customer', 'Year'] + month_names + ['Grand Total']
# some months may be missing in columns; reindex safely
available_cols = [c for c in final_columns if c in pivot_df.columns]
pivot_df = pivot_df.reindex(columns=available_cols)

# store results in session_state for other pages
st.session_state["customer_data_raw"] = df_final
st.session_state["customer_report_pivot"] = pivot_df

# Format pivot_df to excel in memory and offer download
buffer = io.BytesIO()
with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
    pivot_df.to_excel(writer, index=False, sheet_name='Report')
# openpyxl load
buffer.seek(0)
wb = load_workbook(buffer)
ws = wb['Report']

# create col_index_map from current pivot_df
col_index_map = {col: idx + 1 for idx, col in enumerate(pivot_df.columns)}

# apply number format to month columns and Grand Total if present
for col in month_names + ['Grand Total']:
    if col in col_index_map:
        col_idx = col_index_map[col]
        col_letter = get_column_letter(col_idx)
        for cell in ws[col_letter][1:]:
            cell.number_format = '#,##0_);[Red](#,##0)'

# set width for 'Customer' column if present
if 'Customer' in col_index_map:
    ws.column_dimensions[get_column_letter(col_index_map['Customer'])].width = 35

# center-align 'Year' if present
if 'Year' in col_index_map:
    for cell in ws[get_column_letter(col_index_map['Year'])][1:]:
        cell.alignment = Alignment(horizontal='center')

# freeze top row and add autofilter
ws.freeze_panes = 'A2'
ws.auto_filter.ref = ws.dimensions

# save workbook into output buffer
out_buffer = io.BytesIO()
wb.save(out_buffer)
out_buffer.seek(0)

st.success("✅ Customer report generated")
st.markdown("Preview (first 10 rows):")
st.dataframe(pivot_df.head(10))

st.download_button(
    label="📥 Download Customer Monthly Report (Formatted)",
    data=out_buffer.getvalue(),
    file_name="customer_monthly_report.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

# store results in session_state for other pages
st.session_state["customer_data_raw"] = df_final
st.session_state["customer_report_pivot"] = pivot_df

# ✅ for chart_page.py compatibility (like streamlit-official-report)
st.session_state["official_data"] = df_final.copy()
